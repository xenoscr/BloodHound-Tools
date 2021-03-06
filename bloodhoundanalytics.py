#!/usr/bin/python3

import argparse, asyncio, datetime, logging, os, sys, yaml
from openpyxl import Workbook, styles
from openpyxl.utils import get_column_letter
from neo4j import GraphDatabase
from timeit import default_timer as timer

def checkOutputPath(path):
    if os.path.exists(path):
        return True
    else:
        return False

def query_yes_no(question, default="no"):
    valid = {'yes': True, 'y': True, 'ye': True, 'no': False, 'n': False}

    if default == None:
        prompt = ' [y/n] '
    elif default.lower() == 'yes':
        prompt = ' [Y/n] '
    elif default.lower() == 'no':
        prompt = ' [y/N] '
    else:
        raise ValueError('Invalid default answer: {}'.format(default))

    while True:
        sys.stdout.write(question + prompt)
        choice = input().lower()
        if default is not None and choice == '':
            return valid[default]
        elif choice in valid:
            return valid[choice]
        else:
            sys.stdout.write("Please respond with 'yes' or 'no' (or 'y' or 'n').\n")

def newFileName(domain):
    return '{0:%Y%m%d%H%M%s}-{1}.xlsx'.format(datetime.datetime.now(), domain)

def input_default(prompt, default):
    return input("%s [%s] " % (prompt, default)) or default
    
    
def do_statistics_column(bhMain, queryHeading, col, title):
    queries = bhMain.queryData[queryHeading]
    # Run through the queries & updat the cells
    results = []
    for i in range(1, len(queries)):
        result = bhMain.runQuery(queries[i])
        results.append('{name}: {value}'.format(name = queries[i]['name'], value = result))
    bhMain.write_column(0, 1, col, title, results)
    
def do_full_sheet(bhMain, sheetNum, queryHeading):
    queries = bhMain.queryData[queryHeading]
    
    for i in range(1, len(queries)):
        results = bhMain.runQuery(queries[i])
        if results:
            bhMain.write_column(sheetNum, 1, i, queries[i]['name'], results)

def print_header(message):
    fullMsg = '[+] ' + message
    padding = '-' * (80 - 1 - len(fullMsg))
    print(fullMsg + ' ' + padding)
    
def print_prog_info():
    print('-' * 80)
    print('BloodHound Analytics Script\n')
    print('Original Authors:\n\tAndy Robbins (@_wald0),\n\tRohan Vazarkar (@CptJesus)\n\thttps://www.specterops.io/\n')
    print('Python3 Re-Write Author:\n\tConor Richard (@xenoscr)')
    print('-' * 80)
    print('')
    
class bhAnalytics(object):
    def __init__(self, logging, domain, outputFile, queryFile=None):
        # Default Database Values
        self.neo4jDB = 'bolt://localhost:7687'
        self.username = 'neo4j'
        self.password = 'neo4jj'
        self.domain = domain
        self.outputFile = outputFile
        self.getQueryFile(queryFile)
        self.driver = None
        self.db_validated = False
        self.logging = logging

    def getQueryFile(self, queryFile):
        if queryFile is None:
            queryFile = 'queries.yaml'

        with open(queryFile, 'r') as configFile:
            self.queryData = yaml.load(configFile, Loader=yaml.Loader)

    def showDomain(self):
        print('The current domain setting is:')
        print('Domain: {}\n'.format(self.domain))
        
    def updateDomain(self):
        print(self.domain)
        self.domain = input_default('Enter the domain name.', self.domain)
        self.domain = self.domain.upper()
        
    def showDBInfo(self):
        print('The current database settings are:')
        print('Neo4j URL: {}'.format(self.neo4jDB))
        print('Username: {}'.format(self.username))
        print('Password: {}\n'.format(self.password))
    
    def updateDBInfo(self):
        self.neo4jDB = input_default('Enter the Neo4j Bolt URL:', self.neo4jDB)
        self.username = input_default('Enter the username:', self.username)
        self.password = input_default('Enter the password:', self.password)
        print('\n')
    
    def connectDB(self):
        # Close the database if it is connected
        if self.driver is not None:
            self.driver.close()
        
        # Connect to the database
        while self.db_validated == False:
            self.driver = GraphDatabase.driver(self.neo4jDB, auth=(self.username, self.password))
            self.db_validated = self.validateDB()
            if not self.db_validated:
                print('\nUnable to validate provided domain and database settings. Please verify provided values:')
                self.updateDomain()
                self.updateDBInfo()

    def validateDB(self):
        print('Validating Selected Domain')
        session = self.driver.session()
        try:
            result = session.run("MATCH (n {domain:$domain}) RETURN COUNT(n)", domain=self.domain).value()
        except Exception as e:
            self.logging.debug('Unable to connect to the neo4j database')
            self.logging.error(e)
            
        if (int(result[0]) > 0):
            return True
        else:
            return False
    
    def closeDB(self):
        try:
            self.driver.close()
            logging.info('Successfully closed Neo4j database.')
        except Exception as e:
            logging.error('Failed to close Neo4j database.')
            logging.error(e)
    
    def runQuery(self, queryYAML):
        query = queryYAML['query']
        queryType = queryYAML['type']
        try:
            session = self.driver.session()
            start = timer()
            results = session.run(query, domain=self.domain)
            logging.info('{} ran in {}s'.format(queryYAML['name'], timer() - start))
        except Exception as e:
            logging.error('Query failed.')
            logging.error(e)
            raise SystemExit
        if queryType == 'int':
            for result in results:
                return result[0]
        elif queryType == 'list':
            resultList = []
            keys = results.keys()
            for result in results:
                resultList.append(result[keys[0]])
            return resultList
        else:
            return None

    def create_workbook(self):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Front Page'
            wb.create_sheet(title='Critical Assets')
            wb.create_sheet(title='Low Hanging Fruit')
            wb.create_sheet(title='Cross Domain Attacks')
            logging.debug('Workbook successfully initialized.')
            self.workbook = wb
        except Exception as e:
            logging.error('Failed to initialize workbook.')
            logging.error(e)
            raise SystemExit

    def write_cell(self, sheetNum, row, column, text):
        try:
            self.workbook._sheets[sheetNum].cell(row, column, value=text)
        except Exception as e:
            logging.error('Failed to update cell contents.')
            logging.errot(e)

    def write_column(self, sheetNum, row, column, title, results):
        try:
            count = len(results)    
            sheet = self.workbook._sheets[sheetNum]
            # Update title cell
            font = styles.Font(bold=True)
            titleCell = sheet.cell(row, column)
            titleCell.font = font
            sheet.cell(row, column, value=title.format(count))
            
            # Update the rows
            for i in range(0, count):
                sheet.cell(i+row+1, column, value=results[i])
        except Exception as e:
            logging.error('Failed to write column data.')
            logging.error(e)
            
    def save_workbook(self):
        for worksheet in self.workbook._sheets:
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column  # Get the column name
                for cell in col:
                    try:  # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width
        self.workbook.save(outputFile)
        
def main(logging, domain, outputFile):
    # Instantiate a new bhAnalytics Object
    bhMain = bhAnalytics(logging, domain, outputFile)

    # Give a chance to correct the domain
    bhMain.showDomain()
    if query_yes_no('Would you like to change the domain name?'):
        bhMain.updateDomain()
        
    # Show the current DB settings and prompt to change
    bhMain.showDBInfo()
    if query_yes_no('Would you like to change the neo4j DB settings?'):
        IhMain.updateDBInfo()

    # Space it out a bit
    print('\n')
    
    # Connect to the Neo4j Database
    print_header('Connecting to Neo4j Database')
    bhMain.connectDB()

    print_header('Creating workbook')
    # Create a workbook object
    bhMain.create_workbook()
    
    # Front Page -----------------------------------------------------------------------------------------------
    # Node Statistics
    print_header('Running Statistical Analysis')
    do_statistics_column(bhMain, 'node_statistics', 1, 'Node Statistics')
    
    # Edge Statistics
    do_statistics_column(bhMain, 'edge_statistics', 2, 'Edge Statistics')
    
    # QA Info
    do_statistics_column(bhMain, 'qa_statistics', 3, 'QA Information')
    
    # Critical Analysis ----------------------------------------------------------------------------------------
    print_header('Running Critical Asset Analysis')
    do_full_sheet(bhMain, 1, 'critical_asset_analysis')
    
    # Low Hanging Fruit Analysis
    print_header('Running Low Hanging Fruit Analysis')
    do_full_sheet(bhMain, 2, 'low_hanging_fruit')
    
    # Cross Domain Analysis
    print_header('Running Cross Domain Analysis')
    do_full_sheet(bhMain, 3, 'cross_domain_analysis')
    
    # Save the workbook
    print_header('Saving Workbook')
    bhMain.save_workbook()
    
    # Close the Neo4j Database
    bhMain.closeDB()

if __name__ == "__main__":
    print_prog_info()
    # Setup logging
    logLvl = logging.INFO
    logging.basicConfig(level=logLvl, format='%(asctime)s - %(levelname)s: %(message)s')
    logging.debug('Debugging logging is on.')

    # Parse the command line arguments and display a help message in incorrec tparameters are provided
    parser = argparse.ArgumentParser(description = 'Generate a report in spreadsheet format from a neo4j database containing BloodHound data.')
    parser.add_argument('-d', '--domain', type=str, help='The name of the domain that you wish to generate reporting for.')
    parser.add_argument('-o', '--outputfile', type=str, help='The path and name of the file that you wish the report to be written to.')
    args = parser.parse_args()
    
    # Check if the domain name was specified, print help if not
    if args.domain:
        domain = args.domain.upper()
        # If a output file name was specified, check to see if it exists
        
        if args.outputfile:
            if checkOutputPath(args.outputfile):
                answer = query_yes_no('The specified output file already exists. Do you wish to overwrite it?')
                if answer == True:
                    outputFile = args.outputfile
                else:
                    print('Please specify another output file name.')
                    parser.print_help(sys.stderr)
            else:
                outputFile = args.outputfile
        else:
            outputFile = newFileName(domain)
        # Call the main function
        main(logging, domain, outputFile)
    else:
        parser.print_help(sys.stderr)
