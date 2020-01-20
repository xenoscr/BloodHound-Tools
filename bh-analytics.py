#!/usr/bin/python3

import argparse, asyncio, datetime, logging, os, sys, yaml
from openpyxl import Workbook, styles
from openpyxl.utils import get_column_letter
from neo4j import GraphDatabase
from timeit import default_timer as timer

def checkOutputPath(path):
    if os.path.exists(path):
        return True
    else:
        return False

def query_yes_no(question, default="no"):
    valid = {'yes': True, 'y': True, 'ye': True, 'no': False, 'n': False}

    if default == None:
        prompt = ' [y/n] '
    elif default.lower() == 'yes':
        prompt = ' [Y/n] '
    elif default.lower() == 'no':
        prompt = ' [y/N] '
    else:
        raise ValueError('Invalid default answer: {}'.format(default))

    while True:
        sys.stdout.write(question + prompt)
        choice = input().lower()
        if default is not None and choice == '':
            return valid[default]
        elif choice in valid:
            return valid[choice]
        else:
            sys.stdout.write("Please respond with 'yes' or 'no' (or 'y' or 'n').\n")

def newFileName(domain):
    return '{0:%Y%m%d%H%M%s}-{1}.xlsx'.format(datetime.datetime.now(), domain)

def input_default(prompt, default):
    return input("%s [%s] " % (prompt, default)) or default
    
def create_workbook():
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Front Page'
        wb.create_sheet(title='Critical Assets')
        wb.create_sheet(title='Low Hanging Fruit')
        wb.create_sheet(title='Cross Domain Attacks')
        logging.debug('Workbook successfully initialized.')
        return wb
    except Exception as e:
        logging.error('Failed to initialize workbook.')
        logging.error(e)
        raise SystemExit

def write_cell(workbook, sheetNum, row, column, text):
    try:
        workbook._sheets[sheetNum].cell(row, column, value=text)
    except Exception as e:
        logging.error('Failed to update cell contents.')
        logging.errot(e)

def write_column(workbook, sheetNum, row, column, title, results):
    try:
        count = len(results)    
        sheet = workbook._sheets[sheetNum]
        # Update title cell
        font = styles.font(bold=True)
        titleCell = sheet.cell(row, column)
        titleCell.font = font
        sheet.cell(row, column, value=title.format(count))
        
        # Update the rows
        for i in xrange(1, count):
            sheet.cell(i+1, column, value=results[i])
    except Exception as e:
        logging.error('Failed to write column data.')
        logging.error(e)
        
def save_workbook(workbook, outputFile):
    for worksheet in workbook._sheets:
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width
    workbook.save(outputFile)
    
def do_statistics_column(bhMain, workbook, nodeQueries, col, title):
    write_cell(workbook, 0, 1, col, title)
    
    # Run through the queries & updat the cells
    for i in range(1, len(nodeQueries)):
        result = bhMain.runQuery(nodeQueries[i])
        write_cell(workbook, 0, i+1, col, '{name}: {value}'.format(name = nodeQueries[i]['name'], value = result))
    
class bhAnalytics(object):
    def __init__(self, logging, domain, outputFile, queryFile=None):
        # Default Database Values
        self.neo4jDB = 'bolt://localhost:7687'
        self.username = 'neo4j'
        self.password = 'neo4jj'
        self.domain = domain
        self.outputFile = outputFile
        self.getQueryFile(queryFile)
        self.driver = None
        self.db_validated = False
        self.logging = logging

    def getQueryFile(self, queryFile):
        if queryFile is None:
            queryFile = 'queries.yaml'

        with open(queryFile, 'r') as configFile:
            self.queryData = yaml.load(configFile, Loader=yaml.Loader)

    def showDomain(self):
        print('The current domain setting is:')
        print('Domain: {}\n'.format(self.domain))
        
    def updateDomain(self):
        print(self.domain)
        self.domain = input_default('Enter the domain name.', self.domain)
        self.domain = self.domain.upper()
        
    def showDBInfo(self):
        print('The current database settings are:')
        print('Neo4j URL: {}'.format(self.neo4jDB))
        print('Username: {}'.format(self.username))
        print('Password: {}\n'.format(self.password))
    
    def updateDBInfo(self):
        self.neo4jDB = input_default('Enter the Neo4j Bolt URL:', self.neo4jDB)
        self.username = input_default('Enter the username:', self.username)
        self.password = input_default('Enter the password:', self.password)
        print('\n')
    
    def connectDB(self):
        # Close the database if it is connected
        if self.driver is not None:
            self.driver.close()
        
        # Connect to the database
        while self.db_validated == False:
            self.driver = GraphDatabase.driver(self.neo4jDB, auth=(self.username, self.password))
            self.db_validated = self.validateDB()
            if not self.db_validated:
                print('\nUnable to validate provided domain and database settings. Please verify provided values:')
                self.updateDomain()
                self.updateDBInfo()

    def validateDB(self):
        print('Validating Selected Domain')
        session = self.driver.session()
        try:
            result = session.run("MATCH (n {domain:$domain}) RETURN COUNT(n)", domain=self.domain).value()
        except Exception as e:
            self.logging.debug('Unable to connect to the neo4j database')
            self.logging.error(e)
            
        print(result[0])

        if (int(result[0]) > 0):
            return True
        else:
            return False
    
    def closeDB(self):
        try:
            self.driver.close()
            logging.debug('Successfully closed Neo4j database.')
        except Exception as e:
            logging.error('Failed to close Neo4j database.')
            logging.error(e)
    
    def runQuery(self, queryYAML):
        query = queryYAML['query']
        queryType = queryYAML['type']
        try:
            session = self.driver.session()
            results = session.run(query, domain=self.domain)
        except Exception as e:
            logging.error('Query failed.')
            logging.error(e)
            raise SystemExit
        if queryType == 'int':
            for result in results:
                return result[0]
        elif queryType == 'list':
            resultList = []
            keys = results.keys()
            for result in results:
                resultList.append(result[keys[0]])
            return resultList
        else:
            return None

def main(logging, domain, outputFile):
    # Instantiate a new bhAnalytics Object
    bhMain = bhAnalytics(logging, domain, outputFile)

    # Give a chance to correct the domain
    bhMain.showDomain()
    if query_yes_no('Would you like to change the domain name?'):
        bhMain.updateDomain()
        
    # Show the current DB settings and prompt to change
    bhMain.showDBInfo()
    if query_yes_no('Would you like to change the neo4j DB settings?'):
        bhMain.updateDBInfo()

    # Connect to the Neo4j Database
    bhMain.connectDB()

    # Create a workbook object
    bhWorkBook = create_workbook()
    
    # Front Page -----------------------------------------------------------------------------------------------
    # Node Statistics
    do_statistics_column(bhMain, bhWorkBook, bhMain.queryData['front']['node_statistics'], 1, 'Node Statistics')
    # Edge Statistics
    do_statistics_column(bhMain, bhWorkBook, bhMain.queryData['front']['edge_statistics'], 2, 'Edge Statistics')
    # QA Info
    do_statistics_column(bhMain, bhWorkBook, bhMain.queryData['front']['qa_statistics'], 3, 'QA Information')
    
    # Critical Analysis ----------------------------------------------------------------------------------------
    
    # Low Hanging Fruit Analysis
    
    # Cross Domain Analysis
    
    # Save the workbook
    save_workbook(bhWorkBook, outputFile)
    
    # Close the Neo4j Database
    bhMain.closeDB()

if __name__ == "__main__":
    # Setup logging
    logLvl = logging.DEBUG
    logging.basicConfig(level=logLvl, format='%(asctime)s - %(levelname)s - %(message)s')
    logging.debug('Debugging logging is on.')

    # Parse the command line arguments and display a help message in incorrec tparameters are provided
    parser = argparse.ArgumentParser(description = 'Generate a report in spreadsheet format from a neo4j database containing BloodHound data.')
    parser.add_argument('-d', '--domain', type=str, help='The name of the domain that you wish to generate reporting for.')
    parser.add_argument('-o', '--outputfile', type=str, help='The path and name of the file that you wish the report to be written to.')
    args = parser.parse_args()
    
    # Check if the domain name was specified, print help if not
    if args.domain:
        domain = args.domain.upper()
        # If a output file name was specified, check to see if it exists
        
        if args.outputfile:
            if checkOutputPath(args.outputfile):
                answer = query_yes_no('The specified output file already exists. Do you wish to overwrite it?')
                if answer == True:
                    outputFile = args.outputFile
                else:
                    print('Please specify another output file name.')
                    parser.print_help(sys.stderr)
        else:
            outputFile = newFileName(domain)
        # Call the main function
        main(logging, domain, outputFile)
    else:
        parser.print_help(sys.stderr)
